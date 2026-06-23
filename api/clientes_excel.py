"""Exportación e importación de clientes en formato Excel (.xlsx)."""

from __future__ import annotations

import io
import re
import unicodedata
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .models import DIA_SEMANA_COBRANZA_CHOICES, Cliente

DIAS_VALIDOS = {choice[0] for choice in DIA_SEMANA_COBRANZA_CHOICES}

CLIENTE_EXCEL_COLUMNS: tuple[tuple[str, str], ...] = (
    ('nombre', 'Nombre *'),
    ('dni', 'DNI *'),
    ('telefono', 'Teléfono'),
    ('dia_cobro_semanal', 'Día cobro semanal'),
    ('direccion_residencia', 'Dirección residencia'),
    ('direccion_negocio', 'Dirección negocio'),
    ('referencia_parentesco', 'Parentesco referencia'),
    ('referencia_telefono', 'Teléfono referencia'),
    ('referencia', 'Notas referencia'),
    ('actividad_economica', 'Actividad económica'),
)

EJEMPLO_FILA = (
    'Juan Pérez López',
    '0801-1990-12345',
    '9999-8888',
    'lunes',
    'Col. Centro, Tegucigalpa',
    'Mercado central local 12',
    'Hermano',
    '9888-7777',
    'Referencia personal verificada',
    'Comercio minorista',
)

HEADER_ALIASES: dict[str, str] = {
    'nombre': 'nombre',
    'name': 'nombre',
    'dni': 'dni',
    'identidad': 'dni',
    'documento': 'dni',
    'telefono': 'telefono',
    'teléfono': 'telefono',
    'tel': 'telefono',
    'dia_cobro_semanal': 'dia_cobro_semanal',
    'dia cobro semanal': 'dia_cobro_semanal',
    'dia de cobro semanal': 'dia_cobro_semanal',
    'dia_cobro': 'dia_cobro_semanal',
    'direccion_residencia': 'direccion_residencia',
    'dirección residencia': 'direccion_residencia',
    'dir residencia': 'direccion_residencia',
    'direccion_negocio': 'direccion_negocio',
    'dirección negocio': 'direccion_negocio',
    'dir negocio': 'direccion_negocio',
    'referencia_parentesco': 'referencia_parentesco',
    'parentesco referencia': 'referencia_parentesco',
    'parentesco': 'referencia_parentesco',
    'referencia_telefono': 'referencia_telefono',
    'teléfono referencia': 'referencia_telefono',
    'tel referencia': 'referencia_telefono',
    'referencia': 'referencia',
    'notas referencia': 'referencia',
    'notas de referencia': 'referencia',
    'actividad_economica': 'actividad_economica',
    'actividad económica': 'actividad_economica',
    'actividad': 'actividad_economica',
}


def _slug_header(value: Any) -> str:
    if value is None:
        return ''
    text = str(value).strip().lower()
    text = unicodedata.normalize('NFD', text)
    text = ''.join(char for char in text if unicodedata.category(char) != 'Mn')
    text = text.replace('*', '').strip()
    text = re.sub(r'\s+', ' ', text)
    return HEADER_ALIASES.get(text, text.replace(' ', '_'))


def _empty_to_none(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _normalize_dia(value: Any) -> str | None:
    raw = _empty_to_none(value)
    if raw is None:
        return None
    text = raw.lower()
    text = unicodedata.normalize('NFD', text)
    text = ''.join(char for char in text if unicodedata.category(char) != 'Mn')
    if text in DIAS_VALIDOS:
        return text
    raise ValueError(
        f'Día de cobro inválido: "{raw}". Use: lunes, martes, miercoles, jueves, viernes, sabado, domingo.',
    )


def _style_header_row(sheet: Worksheet) -> None:
    header_fill = PatternFill('solid', fgColor='1E3A5F')
    header_font = Font(bold=True, color='FFFFFF')
    for col_idx in range(1, len(CLIENTE_EXCEL_COLUMNS) + 1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font


def _write_workbook(include_example: bool = False) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Clientes'

    for col_idx, (_field, label) in enumerate(CLIENTE_EXCEL_COLUMNS, start=1):
        sheet.cell(row=1, column=col_idx, value=label)
        sheet.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = max(len(label) + 4, 14)

    widths = [28, 18, 14, 18, 32, 32, 20, 18, 28, 28]
    for col_idx, width in enumerate(widths, start=1):
        col_letter = sheet.cell(row=1, column=col_idx).column_letter
        sheet.column_dimensions[col_letter].width = width

    _style_header_row(sheet)

    if include_example:
        for col_idx, value in enumerate(EJEMPLO_FILA, start=1):
            sheet.cell(row=2, column=col_idx, value=value)

    instructions = workbook.create_sheet('Instrucciones')
    instructions['A1'] = 'Campos obligatorios: Nombre, DNI.'
    instructions['A2'] = 'Día cobro semanal: lunes, martes, miercoles, jueves, viernes, sabado, domingo.'
    instructions['A3'] = 'El DNI debe ser único. Filas duplicadas se omiten o actualizan según la opción al importar.'

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def generar_plantilla_clientes_xlsx() -> bytes:
    """Plantilla vacía con encabezados y una fila de ejemplo."""
    return _write_workbook(include_example=True)


def exportar_clientes_xlsx(queryset) -> bytes:
    """Exporta clientes a Excel."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Clientes'

    for col_idx, (_field, label) in enumerate(CLIENTE_EXCEL_COLUMNS, start=1):
        sheet.cell(row=1, column=col_idx, value=label)

    widths = [28, 18, 14, 18, 32, 32, 20, 18, 28, 28]
    for col_idx, width in enumerate(widths, start=1):
        col_letter = sheet.cell(row=1, column=col_idx).column_letter
        sheet.column_dimensions[col_letter].width = width

    _style_header_row(sheet)

    for row_idx, cliente in enumerate(queryset.order_by('nombre'), start=2):
        for col_idx, (field, _label) in enumerate(CLIENTE_EXCEL_COLUMNS, start=1):
            value = getattr(cliente, field, None)
            sheet.cell(row=row_idx, column=col_idx, value=value if value is not None else '')

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _map_headers(sheet: Worksheet) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for col_idx in range(1, sheet.max_column + 1):
        header = _slug_header(sheet.cell(row=1, column=col_idx).value)
        if header in {field for field, _ in CLIENTE_EXCEL_COLUMNS}:
            mapping[col_idx] = header
    return mapping


def _row_is_empty(sheet: Worksheet, row_idx: int, column_map: dict[int, str]) -> bool:
    for col_idx in column_map:
        if _empty_to_none(sheet.cell(row=row_idx, column=col_idx).value) is not None:
            return False
    return True


def _parse_row(sheet: Worksheet, row_idx: int, column_map: dict[int, str]) -> dict[str, Any]:
    data: dict[str, Any] = {}
    for col_idx, field in column_map.items():
        data[field] = sheet.cell(row=row_idx, column=col_idx).value
    return data


def importar_clientes_xlsx(file_obj, *, actualizar_existentes: bool = False) -> dict[str, Any]:
    """Importa clientes desde un archivo .xlsx."""
    try:
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError('El archivo no es un Excel válido (.xlsx).') from exc

    sheet = workbook.active
    if sheet is None or sheet.max_row < 1:
        raise ValueError('El archivo Excel está vacío.')

    column_map = _map_headers(sheet)
    if 'nombre' not in column_map.values() or 'dni' not in column_map.values():
        raise ValueError('El Excel debe incluir las columnas «Nombre» y «DNI» en la primera fila.')

    creados = 0
    actualizados = 0
    omitidos = 0
    errores: list[dict[str, Any]] = []

    for row_idx in range(2, sheet.max_row + 1):
        if _row_is_empty(sheet, row_idx, column_map):
            continue

        raw: dict[str, Any] = {}
        try:
            raw = _parse_row(sheet, row_idx, column_map)
            nombre = _empty_to_none(raw.get('nombre'))
            dni = _empty_to_none(raw.get('dni'))
            if not nombre or not dni:
                raise ValueError('Nombre y DNI son obligatorios.')

            payload = {
                'nombre': nombre[:100],
                'dni': dni[:20],
                'telefono': _empty_to_none(raw.get('telefono')),
                'direccion_residencia': _empty_to_none(raw.get('direccion_residencia')),
                'direccion_negocio': _empty_to_none(raw.get('direccion_negocio')),
                'referencia_parentesco': _empty_to_none(raw.get('referencia_parentesco')),
                'referencia_telefono': _empty_to_none(raw.get('referencia_telefono')),
                'referencia': _empty_to_none(raw.get('referencia')),
                'actividad_economica': _empty_to_none(raw.get('actividad_economica')),
                'dia_cobro_semanal': _normalize_dia(raw.get('dia_cobro_semanal')),
            }

            existente = Cliente.objects.filter(dni__iexact=payload['dni']).first()
            if existente is not None:
                if not actualizar_existentes:
                    omitidos += 1
                    errores.append(
                        {
                            'fila': row_idx,
                            'dni': payload['dni'],
                            'mensaje': 'DNI ya registrado (omitido).',
                        },
                    )
                    continue
                for field, value in payload.items():
                    setattr(existente, field, value)
                existente.save()
                actualizados += 1
            else:
                Cliente.objects.create(**payload)
                creados += 1
        except Exception as exc:
            errores.append(
                {
                    'fila': row_idx,
                    'dni': str(raw.get('dni', '') or ''),
                    'mensaje': str(exc),
                },
            )

    workbook.close()
    return {
        'creados': creados,
        'actualizados': actualizados,
        'omitidos': omitidos,
        'errores': errores,
    }
