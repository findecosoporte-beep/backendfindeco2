"""Exportación Excel y PDF del historial de pagos cobrados."""

from __future__ import annotations

import io
from decimal import Decimal
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from django.utils.dateparse import parse_datetime

from .core.cuotas import extract_cuota_numero_from_documento
from .core.fechas_display import formato_fecha_hora_hn
from .core.findeco_brand import platypus_logo_findeco

MESES_ES = (
    'Enero',
    'Febrero',
    'Marzo',
    'Abril',
    'Mayo',
    'Junio',
    'Julio',
    'Agosto',
    'Septiembre',
    'Octubre',
    'Noviembre',
    'Diciembre',
)

# Mismo orden lógico que el listado en pantalla: cartera → cliente → cobro → montos.
COLUMNAS_EXCEL = (
    ('cartera_nombre', 'Cartera'),
    ('nombre_cliente', 'Cliente'),
    ('dni_cliente', 'DNI'),
    ('numero_prestamo', 'Préstamo'),
    ('documento', 'Documento'),
    ('fecha_programada', 'Fecha programada'),
    ('fecha_pago', 'Fecha canceló'),
    ('registrado_por_nombre', 'Usuario'),
    ('registrado_en', 'Fecha registro'),
    ('capital', 'Capital'),
    ('interes', 'Interés'),
    ('total', 'Total'),
)

PDF_COLUMNAS = (
    ('cartera_nombre', 'Cartera', 22 * mm, True),
    ('nombre_cliente', 'Cliente', 34 * mm, True),
    ('dni_cliente', 'DNI', 26 * mm, True),
    ('numero_prestamo', 'Préstamo', 14 * mm, False),
    ('documento', 'Doc.', 14 * mm, True),
    ('fecha_programada', 'F. programada', 17 * mm, False),
    ('fecha_pago', 'F. canceló', 17 * mm, False),
    ('registrado_por_nombre', 'Usuario', 24 * mm, True),
    ('registrado_en', 'F. registro', 26 * mm, True),
    ('capital', 'Capital', 17 * mm, False),
    ('interes', 'Interés', 15 * mm, False),
    ('total', 'Total', 17 * mm, False),
)


def clave_orden_fila_historial(fila: dict) -> tuple:
    """Orden de impresión: cartera, cliente, fecha programada, cuota, id de pago."""
    fecha_programada = (fila.get('fecha_programada') or '').strip()
    cuota = extract_cuota_numero_from_documento(fila.get('documento'))
    return (
        (fila.get('cartera_nombre') or '').casefold(),
        (fila.get('nombre_cliente') or '').casefold(),
        fecha_programada if fecha_programada else '9999-12-31',
        cuota if cuota is not None else 9999,
        fila.get('id_pago') or 0,
    )


def ordenar_filas_historial(filas: list[dict]) -> None:
    filas.sort(key=clave_orden_fila_historial)


def _filas_ordenadas(datos: dict) -> list[dict]:
    filas = list(datos.get('filas', []))
    ordenar_filas_historial(filas)
    return filas


def _periodo_legible(datos: dict) -> str:
    modo = datos.get('modo', 'dia')
    inicio = datos.get('fecha_inicio', '')
    fin = datos.get('fecha_fin', '')
    if modo == 'dia' and inicio:
        return inicio
    if modo == 'mes' and inicio:
        try:
            y, m, _ = inicio.split('-')
            return f'{MESES_ES[int(m) - 1]} {y}'
        except (ValueError, IndexError):
            return f'{inicio} – {fin}'
    if modo == 'anio' and inicio:
        return inicio[:4]
    if inicio == fin:
        return inicio or '—'
    return f'{inicio} – {fin}'


def nombre_archivo_historial(datos: dict, extension: str) -> str:
    periodo = _periodo_legible(datos).replace(' ', '_').replace('–', '-')
    cartera = (datos.get('cartera_etiqueta') or 'todas').replace(' ', '_')
    return f'historial_pagos_{cartera}_{periodo}.{extension}'


def _formato_generado(iso: str | None) -> str:
    if not iso:
        return '—'
    dt = parse_datetime(iso)
    if dt is None:
        return iso
    return formato_fecha_hora_hn(dt)


def _valor_celda_excel(fila: dict, key: str) -> str:
    val = fila.get(key, '')
    if val not in (None, ''):
        return str(val)
    if key in ('fecha_programada', 'registrado_en', 'registrado_por_nombre', 'documento'):
        return '—'
    return ''


def exportar_historial_pagos_xlsx(datos: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Historial pagos'

    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF')

    ws.append(['FINDECO — Historial de pagos'])
    ws.append([f"Cartera: {datos.get('cartera_etiqueta', 'Todas')}"])
    ws.append([f"Periodo: {_periodo_legible(datos)}"])
    ws.append([f"Generado: {_formato_generado(datos.get('generado_en'))}"])
    ws.append([])

    headers = [label for _, label in COLUMNAS_EXCEL]
    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font

    for fila in _filas_ordenadas(datos):
        ws.append([_valor_celda_excel(fila, key) for key, _ in COLUMNAS_EXCEL])

    resumen = datos.get('resumen', {})
    ws.append([])
    ws.append(
        [
            'Totales',
            '',
            '',
            '',
            '',
            '',
            '',
            '',
            '',
            resumen.get('total_capital', '0'),
            resumen.get('total_interes', '0'),
            resumen.get('total_cobrado', '0'),
        ]
    )
    ws.append(['Registros', resumen.get('registros', 0)])

    for col in ws.columns:
        max_len = 0
        letter_col = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter_col].width = min(max(max_len + 2, 10), 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _money_pdf(value: str | Decimal) -> str:
    try:
        n = Decimal(str(value))
    except (ArithmeticError, ValueError):
        return str(value)
    return f'L {n:,.2f}'


def _pdf_paragraph(text: str | None, style: ParagraphStyle) -> Paragraph:
    raw = (text or '').strip() or '—'
    return Paragraph(escape(raw), style)


def _valor_pdf(fila: dict, key: str, cell_style: ParagraphStyle, wrap: bool):
    if key == 'capital':
        return _money_pdf(fila.get('capital', '0'))
    if key == 'interes':
        return _money_pdf(fila.get('interes', '0'))
    if key == 'total':
        return _money_pdf(fila.get('total', '0'))
    if key == 'fecha_programada':
        text = fila.get('fecha_programada') or '—'
    elif key == 'documento':
        text = fila.get('documento') or '—'
    elif key == 'registrado_por_nombre':
        text = fila.get('registrado_por_nombre') or '—'
    elif key == 'registrado_en':
        text = fila.get('registrado_en') or '—'
    else:
        text = fila.get(key, '') or ''
    if wrap:
        return _pdf_paragraph(str(text), cell_style)
    return str(text or '—')


def exportar_historial_pagos_pdf(datos: dict) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'HistTitle',
        parent=styles['Heading1'],
        fontSize=14,
        alignment=1,
        spaceAfter=6,
    )
    meta_style = ParagraphStyle(
        'HistMeta',
        parent=styles['Normal'],
        fontSize=9,
        alignment=1,
        spaceAfter=2,
    )
    cell_style = ParagraphStyle(
        'HistCell',
        parent=styles['Normal'],
        fontSize=6.5,
        leading=8,
        wordWrap='CJK',
    )

    story = []
    logo = platypus_logo_findeco(ancho_mm=52, alto_mm=20)
    if logo is not None:
        story.extend([logo, Spacer(1, 4)])
    story.extend(
        [
            Paragraph('FINDECO — Historial de pagos', title_style),
            Paragraph(f"Cartera: {datos.get('cartera_etiqueta', 'Todas')}", meta_style),
            Paragraph(f"Periodo: {_periodo_legible(datos)}", meta_style),
            Paragraph(f"Generado: {_formato_generado(datos.get('generado_en'))}", meta_style),
            Spacer(1, 8),
        ]
    )

    headers = [header for _, header, _, _ in PDF_COLUMNAS]
    table_data = [headers]
    for fila in _filas_ordenadas(datos):
        table_data.append(
            [
                _valor_pdf(fila, key, cell_style, wrap)
                for key, _, _, wrap in PDF_COLUMNAS
            ]
        )

    resumen = datos.get('resumen', {})
    totales = ['TOTALES'] + [''] * (len(PDF_COLUMNAS) - 5)
    totales.append(f"{resumen.get('registros', 0)} reg.")
    totales.extend(
        [
            _money_pdf(resumen.get('total_capital', '0')),
            _money_pdf(resumen.get('total_interes', '0')),
            _money_pdf(resumen.get('total_cobrado', '0')),
        ]
    )
    table_data.append(totales)

    col_widths = [width for _, _, width, _ in PDF_COLUMNAS]
    money_col = len(PDF_COLUMNAS) - 3
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('FONTSIZE', (0, 1), (-1, -1), 6.5),
                ('ALIGN', (money_col, 0), (-1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8EEF4')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F7F9FC')]),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    return buffer.getvalue()
